"""
Route files to low-level parsers and return ParsedDocument.

Supports: .txt, .md, .csv, .json, .xlsx (safe v1). Optional: .pdf, .docx.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from workflow_dataset.parse.document_models import (
    ParsedDocument,
    ParsedSection,
    ParsedTable,
    DocumentSignal,
    ExtractionPolicy,
)
from workflow_dataset.parse.artifact_classifier import classify_artifact


def _read_text_safe(path: Path, max_bytes: int = 500_000, encoding: str = "utf-8") -> str:
    try:
        with open(path, "r", encoding=encoding, errors="replace") as f:
            return f.read(max_bytes)
    except OSError:
        return ""


def _parse_txt(path: Path, policy: ExtractionPolicy) -> ParsedDocument:
    text = _read_text_safe(path)
    summary = text[:500].strip() if policy != ExtractionPolicy.METADATA_ONLY else ""
    return ParsedDocument(
        source_path=str(path.resolve()),
        artifact_family="text_document",
        title=path.name,
        summary=summary,
        sections=[ParsedSection(content=text[:2000])] if text and policy in (ExtractionPolicy.SIGNALS_AND_SUMMARIES, ExtractionPolicy.FULL_TEXT) else [],
        metadata={"extension": "txt", "size": path.stat().st_size if path.exists() else 0},
        policy_used=policy,
        raw_text_snippet=text[:2000] if policy == ExtractionPolicy.FULL_TEXT else (text[:300] if policy == ExtractionPolicy.SIGNALS_AND_SUMMARIES else ""),
    )


def _parse_md(path: Path, policy: ExtractionPolicy) -> ParsedDocument:
    text = _read_text_safe(path)
    sections: list[ParsedSection] = []
    current_heading = ""
    current_content: list[str] = []
    for line in text.splitlines():
        if line.startswith("#"):
            if current_content:
                sections.append(ParsedSection(heading=current_heading, content="\n".join(current_content)))
            level = len(line) - len(line.lstrip("#"))
            current_heading = line.lstrip("#").strip()
            current_content = []
            sections.append(ParsedSection(heading=current_heading, content="", level=min(level, 6)))
        else:
            current_content.append(line)
    if current_content and sections:
        sections[-1].content = "\n".join(current_content)
    elif current_content:
        sections.append(ParsedSection(content="\n".join(current_content)))
    summary = text[:500].strip() if policy != ExtractionPolicy.METADATA_ONLY else ""
    return ParsedDocument(
        source_path=str(path.resolve()),
        artifact_family="text_document",
        title=path.name,
        summary=summary,
        sections=sections,
        metadata={"extension": "md", "size": path.stat().st_size if path.exists() else 0},
        policy_used=policy,
        raw_text_snippet=text[:2000] if policy in (ExtractionPolicy.SIGNALS_AND_SUMMARIES, ExtractionPolicy.FULL_TEXT) else "",
    )


def _parse_csv(path: Path, policy: ExtractionPolicy) -> ParsedDocument:
    import csv
    text = _read_text_safe(path, max_bytes=100_000)
    rows: list[list[str]] = []
    try:
        for row in csv.reader(text.splitlines()[:500]):
            rows.append(row)
    except Exception:
        rows = []
    headers = rows[0] if rows else []
    table = ParsedTable(headers=headers, rows=rows[1:21], source_path=str(path)) if rows else ParsedTable()
    summary = f"CSV with {len(rows)} rows" if policy != ExtractionPolicy.METADATA_ONLY else ""
    return ParsedDocument(
        source_path=str(path.resolve()),
        artifact_family="spreadsheet_table",
        title=path.name,
        summary=summary,
        tables=[table],
        signals=[DocumentSignal(signal_type="schema_headers", value=headers, source_path=str(path))] if headers else [],
        metadata={"extension": "csv", "rows": len(rows), "size": path.stat().st_size if path.exists() else 0},
        policy_used=policy,
    )


def _parse_json(path: Path, policy: ExtractionPolicy) -> ParsedDocument:
    import json as _json
    text = _read_text_safe(path, max_bytes=200_000)
    try:
        data = _json.loads(text)
        if isinstance(data, dict):
            keys = list(data.keys())[:20]
            signals = [DocumentSignal(signal_type="top_level_keys", value=keys, source_path=str(path))]
        elif isinstance(data, list):
            signals = [DocumentSignal(signal_type="list_length", value=len(data), source_path=str(path))]
        else:
            signals = []
    except _json.JSONDecodeError:
        signals = []
        data = None
    summary = str(type(data).__name__) + (" keys: " + ", ".join(list(data.keys())[:8]) if isinstance(data, dict) else "") if data and policy != ExtractionPolicy.METADATA_ONLY else ""
    return ParsedDocument(
        source_path=str(path.resolve()),
        artifact_family="text_document",
        title=path.name,
        summary=summary[:500],
        signals=signals,
        metadata={"extension": "json", "size": path.stat().st_size if path.exists() else 0},
        policy_used=policy,
    )


def _parse_xlsx(path: Path, policy: ExtractionPolicy) -> ParsedDocument:
    tables: list[ParsedTable] = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for name in wb.sheetnames[:10]:
            sh = wb[name]
            rows = list(sh.iter_rows(values_only=True))[:100]
            if rows:
                headers = [str(c) for c in rows[0]]
                table_rows = [[str(c) for c in row] for row in rows[1:]]
                tables.append(ParsedTable(headers=headers, rows=table_rows, sheet_name=name, source_path=str(path)))
        wb.close()
    except Exception as e:
        return ParsedDocument(
            source_path=str(path.resolve()),
            artifact_family="spreadsheet_table",
            title=path.name,
            error=str(e),
            policy_used=policy,
        )
    summary = f"Excel with {len(tables)} sheet(s)" if policy != ExtractionPolicy.METADATA_ONLY else ""
    return ParsedDocument(
        source_path=str(path.resolve()),
        artifact_family="spreadsheet_table",
        title=path.name,
        summary=summary,
        tables=tables,
        signals=[DocumentSignal(signal_type="sheet_names", value=[t.sheet_name for t in tables], source_path=str(path))] if tables else [],
        metadata={"extension": "xlsx", "sheets": len(tables), "size": path.stat().st_size if path.exists() else 0},
        policy_used=policy,
    )


def route_and_parse_file(
    path: Path | str,
    policy: ExtractionPolicy = ExtractionPolicy.SIGNALS_AND_SUMMARIES,
    max_file_size: int = 10 * 1024 * 1024,
) -> ParsedDocument:
    """
    Route a file to the appropriate parser. Returns ParsedDocument (or error in .error).
    Safe: respects max_file_size, no network. Supports txt, md, csv, json, xlsx.
    """
    p = Path(path)
    if not p.exists() or p.is_dir():
        return ParsedDocument(source_path=str(p), artifact_family="unknown", error="not a file or missing")
    try:
        if p.stat().st_size > max_file_size:
            return ParsedDocument(source_path=str(p), artifact_family="unknown", error="file too large", metadata={"size": p.stat().st_size})
    except OSError:
        return ParsedDocument(source_path=str(p), artifact_family="unknown", error="stat failed")
    ext = p.suffix.lstrip(".").lower()
    if ext == "txt":
        return _parse_txt(p, policy)
    if ext == "md" or ext == "markdown":
        return _parse_md(p, policy)
    if ext == "csv":
        return _parse_csv(p, policy)
    if ext == "json":
        return _parse_json(p, policy)
    if ext == "xlsx":
        return _parse_xlsx(p, policy)
    # Unsupported: metadata-only placeholder
    family = classify_artifact(p)
    return ParsedDocument(
        source_path=str(p.resolve()),
        artifact_family=family.value,
        title=p.name,
        metadata={"extension": ext, "size": p.stat().st_size},
        policy_used=policy,
        error="no parser for extension",
    )
